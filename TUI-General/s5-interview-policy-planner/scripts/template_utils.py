from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import cell_value_to_text, infer_district, load_json, normalize_string, parse_amount_to_wanyuan, parse_int


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DEFAULT_REGISTRY_PATH = SKILL_DIR / "templates" / "registry.json"


def load_registry(path: Path = DEFAULT_REGISTRY_PATH) -> dict[str, Any]:
    return load_json(path)


def resolve_template_profile(workbook_path: Path, registry_path: Path = DEFAULT_REGISTRY_PATH) -> dict[str, Any]:
    registry = load_registry(registry_path)
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        for template in registry["templates"]:
            match = template["match"]
            sheet_name = match["sheet_name"]
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            title = cell_value_to_text(ws[match["title_cell"]].value) or ""
            if match["title_contains"] not in title:
                continue
            profile_path = registry_path.parent / template["profile"]
            return load_json(profile_path)
    finally:
        workbook.close()
    raise ValueError(f"未识别模板: {workbook_path}")


def extract_seed_context(workbook_path: Path, profile: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        ws = workbook[profile["sheet_name"]]
        seed = {}
        for key, cell in profile.get("seed_cells", {}).items():
            seed[key] = cell_value_to_text(ws[cell].value)

        seed["district"] = infer_district(seed.get("address") or seed.get("industry") or seed.get("enterprise_name"))
        capital = parse_amount_to_wanyuan(seed.get("registered_capital_text"))
        if capital is not None:
            seed["registered_capital_wanyuan"] = capital
        employee = parse_int(seed.get("employee_count_text"))
        if employee is not None:
            seed["employee_count"] = employee
        return seed
    finally:
        workbook.close()


def split_project_names(text: str | None) -> list[str]:
    if not text:
        return []
    parts = re.split(r"[、；;\n]+", str(text))
    return [normalize_string(item) for item in parts if normalize_string(item)]


def scan_sheet_sections(ws, profile: dict[str, Any]) -> list[dict[str, Any]]:
    sections = profile.get("sections", [])
    normalized_sections = [(normalize_string(section["name"]), section) for section in sections]
    discovered: list[dict[str, Any]] = []
    seen: set[str] = set()

    for row_number in range(1, ws.max_row + 1):
        title = normalize_string(cell_value_to_text(ws.cell(row_number, 1).value))
        if not title:
            continue
        matched_name = None
        matched_section = None
        for section_name, section in normalized_sections:
            if section_name in seen:
                continue
            if title == section_name or section_name in title:
                matched_name = section_name
                matched_section = section
                break
        if matched_section is None:
            continue
        seen.add(matched_name or "")
        discovered.append({"section": matched_section, "title_row": row_number})

    if len(discovered) != len(sections):
        missing = [section["name"] for section_name, section in normalized_sections if section_name not in seen]
        raise ValueError(f"模板缺少主表分区标题: {'、'.join(missing)}")

    discovered.sort(key=lambda item: item["title_row"])
    resolved: list[dict[str, Any]] = []
    for index, item in enumerate(discovered):
        section = dict(item["section"])
        title_row = item["title_row"]
        next_title_row = discovered[index + 1]["title_row"] if index + 1 < len(discovered) else ws.max_row + 1
        section["title_row"] = title_row
        section["header_row"] = title_row + 1
        section["first_data_row"] = title_row + 2
        section["next_section_title_row"] = next_title_row
        section["data_end_row"] = next_title_row - 1
        resolved.append(section)
    return resolved


def extract_existing_project_names_from_sheet(ws, profile: dict[str, Any]) -> set[str]:
    names: set[str] = set()
    applied_projects_cell = profile.get("seed_cells", {}).get("applied_projects")
    if applied_projects_cell:
        for item in split_project_names(cell_value_to_text(ws[applied_projects_cell].value)):
            names.add(item)

    for section in scan_sheet_sections(ws, profile):
        project_col = section["columns"]["项目名称"]
        for row_number in range(section["first_data_row"], section["data_end_row"] + 1):
            value = cell_value_to_text(ws.cell(row_number, project_col).value)
            normalized = normalize_string(value) if value else ""
            if normalized:
                names.add(normalized)
    return names


def extract_existing_project_names(workbook_path: Path, profile: dict[str, Any]) -> set[str]:
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        ws = workbook[profile["sheet_name"]]
        return extract_existing_project_names_from_sheet(ws, profile)
    finally:
        workbook.close()


def section_capacity(section: dict[str, Any]) -> int:
    return len(section.get("row_slots", []))


def row_pair_for_index(section: dict[str, Any], index: int) -> tuple[int, int | None]:
    slots = section.get("row_slots", [])
    data_row = slots[index] if index < len(slots) else section["first_data_row"] + index * 2
    spacer_row = data_row + 1
    return data_row, spacer_row
