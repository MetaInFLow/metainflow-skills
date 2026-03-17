from __future__ import annotations

import argparse
import json
import locale
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
WORKSPACE_ROOT = SKILL_DIR.parent
DEFAULT_CONFIG_PATH = SKILL_DIR / "evals" / "evals.json"
DEFAULT_OUTPUT_ROOT = SKILL_DIR / "output" / "s5-interview-policy-planner-evals"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the local eval suite for interview-policy-planner.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    return parser.parse_args()


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def resolve_path(path_text: str) -> Path:
    path = Path(path_text)
    return path if path.is_absolute() else (SKILL_DIR / path).resolve()


def decode_output(raw: bytes) -> str:
    return raw.decode(locale.getpreferredencoding(False) or "utf-8", errors="replace").strip()


def read_review_sheet_info(workbook_path: Path | None) -> dict[str, Any]:
    if workbook_path is None or not workbook_path.exists():
        return {"updated_workbook_present": False, "review_sheet_present": False, "review_sheet_rows": 0}
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        has_review = "审核与补采" in workbook.sheetnames
        return {
            "updated_workbook_present": True,
            "review_sheet_present": has_review,
            "review_sheet_rows": workbook["审核与补采"].max_row if has_review else 0,
        }
    finally:
        workbook.close()


def collect_review_counts(matches_payload: dict[str, Any]) -> dict[str, int]:
    reviewable = [item for item in matches_payload.get("matches", []) if item.get("fit_status") in ("符合", "有条件符合", "证据不足")]
    return {
        "selected_policy_count": len(reviewable),
        "eligible_count": sum(1 for item in reviewable if item.get("fit_status") == "符合"),
        "conditional_count": sum(1 for item in reviewable if item.get("fit_status") == "有条件符合"),
        "insufficient_count": sum(1 for item in reviewable if item.get("fit_status") == "证据不足"),
    }


def inspect_outputs(output_dir: Path) -> dict[str, Any]:
    result: dict[str, Any] = {}
    profile_path = output_dir / "company_profile.json"
    matches_path = output_dir / "policy_matches.json"
    session_state_path = output_dir / "session_state.json"
    preview_path = output_dir / "preview_candidates.json"
    workbook_path = next(output_dir.glob("*_updated.xlsx"), None)

    if profile_path.exists():
        result["profile"] = load_json(profile_path)
    if matches_path.exists():
        result["matches"] = load_json(matches_path)
        result.update(collect_review_counts(result["matches"]))
    if session_state_path.exists():
        result["session_state"] = load_json(session_state_path)
        result["session_status"] = result["session_state"].get("status")
        result["pending_question_count"] = len(result["session_state"].get("pending_questions", []))
        result["gating_reason_count"] = len(result["session_state"].get("gating_reasons", []))
    if preview_path.exists():
        preview = load_json(preview_path)
        result["preview_candidates"] = preview
        result["preview_candidate_count"] = len(preview)
    result.update(read_review_sheet_info(workbook_path))
    return result


def append_failure(failures: list[str], name: str, expected: Any, actual: Any) -> None:
    failures.append(f"{name}: expected {expected!r}, got {actual!r}")


def check_expectations(expect: dict[str, Any], inspected: dict[str, Any], failures: list[str]) -> None:
    if "updated_workbook_present" in expect and inspected.get("updated_workbook_present") != expect["updated_workbook_present"]:
        append_failure(failures, "updated_workbook_present", expect["updated_workbook_present"], inspected.get("updated_workbook_present"))
    if "review_sheet_present" in expect and inspected.get("review_sheet_present") != expect["review_sheet_present"]:
        append_failure(failures, "review_sheet_present", expect["review_sheet_present"], inspected.get("review_sheet_present"))
    if "session_status" in expect and inspected.get("session_status") != expect["session_status"]:
        append_failure(failures, "session_status", expect["session_status"], inspected.get("session_status"))
    if "pending_question_count" in expect and inspected.get("pending_question_count") != expect["pending_question_count"]:
        append_failure(failures, "pending_question_count", expect["pending_question_count"], inspected.get("pending_question_count"))
    if "preview_candidate_count" in expect and inspected.get("preview_candidate_count") != expect["preview_candidate_count"]:
        append_failure(failures, "preview_candidate_count", expect["preview_candidate_count"], inspected.get("preview_candidate_count"))
    for field_name, expected_value in expect.get("transcript_facts", {}).items():
        actual_value = inspected.get("profile", {}).get("transcript_facts", {}).get(field_name, {}).get("value")
        if actual_value != expected_value:
            append_failure(failures, f"transcript_facts.{field_name}", expected_value, actual_value)
    for key, expected_value in expect.get("review_counts", {}).items():
        actual_value = inspected.get(key)
        if actual_value != expected_value:
            append_failure(failures, key, expected_value, actual_value)


def build_step_command(case: dict[str, Any], step: dict[str, Any], output_dir: Path, suite_today: str | None) -> tuple[list[str], Path | None]:
    command = [
        sys.executable,
        str(SCRIPT_DIR / "run_pipeline.py"),
        "--output-dir",
        str(output_dir),
        "--stage",
        step.get("stage") or case.get("stage") or "full",
    ]
    workbook = step.get("workbook") or case.get("workbook")
    minutes = step.get("minutes") or case.get("minutes")
    if workbook:
        command.extend(["--workbook", str(resolve_path(workbook))])
    if minutes:
        command.extend(["--minutes", str(resolve_path(minutes))])
    if case.get("policy_csv"):
        command.extend(["--policy-csv", str(resolve_path(case["policy_csv"]))])
    today_value = step.get("today") or case.get("today") or suite_today
    if today_value:
        command.extend(["--today", today_value])
    confirmations_path = None
    if step.get("confirmations") is not None:
        confirmations_path = output_dir / f"{step.get('id', 'step')}_confirmations.json"
        write_json(confirmations_path, step["confirmations"])
        command.extend(["--confirmations", str(confirmations_path)])
    return command, confirmations_path


def run_step(case: dict[str, Any], step: dict[str, Any], output_dir: Path, suite_today: str | None) -> dict[str, Any]:
    command, confirmations_path = build_step_command(case, step, output_dir, suite_today)
    started_at = time.perf_counter()
    completed = subprocess.run(command, capture_output=True, text=False)
    duration_seconds = round(time.perf_counter() - started_at, 2)
    result = {
        "id": step.get("id") or step.get("stage") or "step",
        "exit_code": completed.returncode,
        "duration_seconds": duration_seconds,
        "stdout": decode_output(completed.stdout),
        "stderr": decode_output(completed.stderr),
        "failures": [],
        "passed": True,
    }
    expect = step.get("expect", {})
    success = completed.returncode == 0
    if success != expect.get("success", True):
        append_failure(result["failures"], "success", expect.get("success", True), success)
    if not success:
        if "error_contains" in expect and expect["error_contains"] not in result["stderr"]:
            result["failures"].append(f"error_contains: expected substring {expect['error_contains']!r}")
        result["passed"] = not result["failures"]
        return result

    inspected = inspect_outputs(output_dir)
    result.update({key: value for key, value in inspected.items() if key not in {"profile", "matches", "session_state", "preview_candidates"}})
    check_expectations(expect, inspected, result["failures"])
    result["passed"] = not result["failures"]
    if confirmations_path and confirmations_path.exists():
        confirmations_path.unlink()
    return result


def evaluate_case(case: dict[str, Any], output_root: Path, suite_today: str | None) -> dict[str, Any]:
    case_id = case["id"]
    output_dir = output_root / case_id
    if output_dir.exists():
        try:
            shutil.rmtree(output_dir)
        except PermissionError:
            suffix = int(time.time() * 1000)
            output_dir = output_root / f"{case_id}_{suffix}"
    output_dir.mkdir(parents=True, exist_ok=True)

    steps = case.get("steps") or [{"id": case_id, "stage": case.get("stage") or "full", "expect": case.get("expect", {})}]
    step_results = [run_step(case, step, output_dir, suite_today) for step in steps]
    return {
        "id": case_id,
        "output_dir": str(output_dir),
        "passed": all(step["passed"] for step in step_results),
        "failures": [failure for step in step_results for failure in step["failures"]],
        "steps": step_results,
    }


def main() -> int:
    args = parse_args()
    config = load_json(args.config)
    results = [evaluate_case(case, args.output_root, config.get("today")) for case in config.get("evals", [])]
    payload = {
        "skill_name": config.get("skill_name", "s5-interview-policy-planner"),
        "today": config.get("today"),
        "passed": sum(1 for item in results if item["passed"]),
        "failed": sum(1 for item in results if not item["passed"]),
        "results": results,
    }
    write_json(args.output_root / "summary.json", payload)
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if payload["failed"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
