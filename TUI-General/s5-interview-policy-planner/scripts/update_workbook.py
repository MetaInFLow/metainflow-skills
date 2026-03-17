from __future__ import annotations

import argparse
import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from common import (
    estimate_row_height,
    format_policy_conditions,
    FIT_STATUS_CONDITIONAL,
    FIT_STATUS_ELIGIBLE,
    FIT_STATUS_INSUFFICIENT_EVIDENCE,
    REVIEWABLE_FIT_STATUSES,
    clean_text,
    copy_row_style,
    normalize_fit_status,
    normalize_string,
    parse_int,
    stringify_value,
)
from template_utils import (
    extract_existing_project_names_from_sheet,
    extract_seed_context,
    resolve_template_profile,
    scan_sheet_sections,
)


MISSING_GUIDE = {
    "region": ("用于过滤区级、市级、省级政策。", "确认注册地或实际经营地所属区。", "所有区级政策"),
    "annual_output_wanyuan": ("用于判断营收门槛和资金强度相关政策。", "补近一年真实营收或开票销售额。", "营收门槛类政策"),
    "rd_ratio_pct": ("用于判断研发投入门槛。", "补研发费用占营业收入比例。", "研发投入类政策"),
    "patent_count_total": ("用于判断知识产权门槛。", "补已授权知识产权总量。", "知识产权类政策"),
    "patent_count_invention": ("用于判断发明专利类门槛。", "补已授权发明专利数量。", "发明专利门槛类政策"),
    "rd_staff_count": ("用于判断研发团队要求。", "补专职研发人员数量。", "研发团队类政策"),
    "main_product": ("用于判断产业方向和项目相关性。", "补当前主要产品或主营业务。", "行业定向政策"),
    "high_tech_enterprise": ("部分政策需要高企身份或会把其作为加分项。", "确认是否已取得高新技术企业资质。", "高企相关政策"),
}

HIGH_PRIORITY_FIELDS = {
    "region",
    "annual_output_wanyuan",
    "patent_count_total",
    "rd_ratio_pct",
}

FIELD_LABELS = {
    "high_tech_enterprise": "高新技术企业状态",
    "contact_person": "联系人",
    "contact_email": "联系邮箱",
    "contact_phone": "联系电话",
    "patent_count_invention": "发明专利数量",
    "patent_count_utility_model": "实用新型专利数量",
    "patent_count_total": "知识产权数量",
    "rd_ratio_pct": "研发投入占比",
    "rd_staff_count": "专职研发人数",
    "annual_output_wanyuan": "营业收入/年产值",
    "main_product": "主要产品",
    "region": "所属地区",
}

TOP_CONTEXT_NOTE = ("工作簿顶部基础信息缺失，会影响政策判断。", "请先在目标 Excel 顶部补充该字段。", "匹配基础画像")
MAIN_SHEET_MAX_COL = 8
AI_SEQUENCE_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
STATUS_PRIORITY = {
    FIT_STATUS_ELIGIBLE: 0,
    FIT_STATUS_CONDITIONAL: 1,
    FIT_STATUS_INSUFFICIENT_EVIDENCE: 2,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update a workbook with policy matches and add a review sheet.")
    parser.add_argument("--workbook", type=Path, required=True, help="Target workbook path.")
    parser.add_argument("--matches", type=Path, required=True, help="policy_matches.json path.")
    parser.add_argument("--profile", type=Path, required=True, help="company_profile.json path.")
    parser.add_argument("--registry", type=Path, required=True, help="template registry path.")
    parser.add_argument("--output", type=Path, required=True, help="Output workbook path.")
    return parser.parse_args()

def build_missing_rows(profile: dict[str, Any], template_profile: dict[str, Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    top_fields = {item["name"]: item for item in template_profile.get("top_fields", [])}
    ordered_missing = [name for name in top_fields if name in profile.get("missing_fields", [])]
    ordered_missing.extend(name for name in profile.get("missing_fields", []) if name not in top_fields)

    for field_name in ordered_missing:
        if field_name in top_fields:
            why, ask, impact = TOP_CONTEXT_NOTE
        else:
            why, ask, impact = MISSING_GUIDE.get(field_name, ("用于继续判断政策适配性。", "补充该字段的真实值或证明材料。", "审核页候选政策"))
        label = top_fields.get(field_name, {}).get("label", FIELD_LABELS.get(field_name, field_name))
        priority = "高" if field_name in HIGH_PRIORITY_FIELDS or field_name in top_fields else "中"
        rows.append([label, why, ask, impact, priority])
    if not rows:
        rows.append(["无", "当前输入已覆盖主要匹配字段。", "如需提高准确率，可补充证明材料。", "低风险候选政策复核", "低"])
    return rows


def build_summary_rows(summary: dict[str, Any]) -> list[list[str]]:
    rows = []
    for item in summary.get("items", []):
        rows.append([item.get("label", "摘要"), item.get("value", "")])
    if not rows:
        rows.append(["摘要", summary.get("text", "无")])
    return rows


def write_summary_section(ws, row_number: int, title: str, summary: dict[str, Any]) -> int:
    ws.cell(row_number, 1).value = title
    ws.cell(row_number, 1).font = Font(bold=True)
    row_number += 1
    ws.cell(row_number, 1).value = "维度"
    ws.cell(row_number, 2).value = "内容"
    ws.cell(row_number, 1).font = Font(bold=True)
    ws.cell(row_number, 2).font = Font(bold=True)
    row_number += 1
    for label, value in build_summary_rows(summary):
        ws.cell(row_number, 1).value = label
        ws.cell(row_number, 2).value = value
        ws.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
        row_number += 1
    return row_number


def build_conflict_rows(profile: dict[str, Any], template_profile: dict[str, Any]) -> list[list[str]]:
    top_fields = {item["name"]: item for item in template_profile.get("top_fields", [])}
    rows: list[list[str]] = []
    for item in profile.get("conflicts", []):
        field_name = item.get("field") or ""
        label = top_fields.get(field_name, {}).get("label", FIELD_LABELS.get(field_name, field_name))
        rows.append(
            [
                label,
                stringify_value(item.get("excel_value")),
                stringify_value(item.get("transcript_value")),
                "以 Excel 顶部字段为准，不自动改写",
                item.get("reason") or "",
            ]
        )
    if not rows:
        rows.append(["无", "", "", "无需处理", "未发现顶部字段与访谈纪要冲突"])
    return rows


def build_policy_result_rows(matches: dict[str, Any]) -> list[list[str]]:
    rows = []
    for item in matches.get("matches", []):
        fit_status = normalize_fit_status(item.get("fit_status"))
        if fit_status not in REVIEWABLE_FIT_STATUSES:
            continue
        evidence_and_gaps = list(dict.fromkeys([*(item.get("missing_evidence", []) or []), *(item.get("gap_clauses", []) or [])]))
        rows.append(
            [
                item.get("section_name") or item.get("section_hint") or "",
                item.get("项目名称") or "",
                fit_status,
                item.get("reason") or "",
                "、".join(evidence_and_gaps),
                stringify_value(item.get("source_row")),
            ]
        )
    if not rows:
        rows.append(["无审核候选", "", "", "当前未匹配到“符合”“有条件符合”或“证据不足”的政策。", "", ""])
    return rows


def fallback_section_hint(match: dict[str, Any]) -> str | None:
    project_type = clean_text(match.get("项目类型", ""))
    support_mode = clean_text(match.get("支持方式", ""))
    title = clean_text(match.get("项目名称", ""))
    strength = clean_text(match.get("资助金额", ""))
    text = " ".join([project_type, support_mode, title, strength])

    if "资质" in project_type or "认定" in project_type:
        return "资质认定类"
    if any(token in project_type for token in ("资助", "扶持", "奖补", "奖励")):
        return "资助类"
    if "认定" in project_type or "认定" in title:
        return "资质认定类"
    if any(token in text for token in ("资助", "补助", "奖补", "扶持", "配套")):
        return "资助类"
    if any(token in text for token in ("评选", "评价", "征集", "示范", "中心", "品牌", "小巨人", "单项冠军", "专利奖", "领军企业", "创新产品")):
        return "资质认定类"
    return None


def resolve_main_sheet_section_name(match: dict[str, Any], template_profile: dict[str, Any]) -> str | None:
    sections = template_profile.get("sections", [])
    name_to_name = {section["name"]: section["name"] for section in sections}
    hint_to_name = {section["section_hint"]: section["name"] for section in sections}

    section_name = match.get("section_name")
    if section_name in name_to_name:
        return section_name

    section_hint = match.get("section_hint")
    if section_hint in hint_to_name:
        return hint_to_name[section_hint]

    fallback_hint = fallback_section_hint(match)
    if fallback_hint in hint_to_name:
        return hint_to_name[fallback_hint]
    return None


def sort_review_candidates(matches: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        matches,
        key=lambda item: (
            STATUS_PRIORITY.get(normalize_fit_status(item.get("fit_status")), 99),
            -(item.get("fit_score") or 0.0),
            clean_text(item.get("项目名称", "")),
        ),
    )


def build_main_sheet_candidates(matches: dict[str, Any], template_profile: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    grouped = {section["name"]: [] for section in template_profile.get("sections", [])}
    for item in matches.get("matches", []):
        fit_status = normalize_fit_status(item.get("fit_status"))
        if fit_status not in REVIEWABLE_FIT_STATUSES or not item.get("review_candidate"):
            continue
        section_name = resolve_main_sheet_section_name(item, template_profile)
        if section_name is None:
            continue
        grouped.setdefault(section_name, []).append(item)
    return {name: sort_review_candidates(items) for name, items in grouped.items()}


def is_blank_main_sheet_row(ws, row_number: int, max_col: int = MAIN_SHEET_MAX_COL) -> bool:
    for col in range(1, max_col + 1):
        value = ws.cell(row_number, col).value
        if clean_text(str(value)) if value is not None else "":
            return False
    return True


def row_has_main_sheet_content(ws, row_number: int, max_col: int = MAIN_SHEET_MAX_COL) -> bool:
    return not is_blank_main_sheet_row(ws, row_number, max_col=max_col)


def section_used_rows(ws, section: dict[str, Any]) -> list[int]:
    return [
        row_number
        for row_number in range(section["first_data_row"], section["data_end_row"] + 1)
        if row_has_main_sheet_content(ws, row_number)
    ]


def max_section_sequence(ws, section: dict[str, Any]) -> int:
    sequence_col = section["columns"]["序号"]
    max_value = 0
    for row_number in section_used_rows(ws, section):
        value = parse_int(ws.cell(row_number, sequence_col).value)
        if value is not None:
            max_value = max(max_value, value)
    return max_value


def shift_merged_ranges(ws, row_number: int, amount: int) -> None:
    if amount <= 0:
        return
    shifted_row_dimensions = {index: copy(dimension) for index, dimension in ws.row_dimensions.items() if index >= row_number}
    shifted: list[tuple[int, int, int, int]] = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= row_number:
            shifted.append((merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col))
            ws.unmerge_cells(str(merged_range))
    ws.insert_rows(row_number, amount=amount)
    for index in sorted(shifted_row_dimensions):
        dimension = copy(shifted_row_dimensions[index])
        dimension.index = index + amount
        ws.row_dimensions[index + amount] = dimension
    for min_row, max_row, min_col, max_col in shifted:
        ws.merge_cells(
            start_row=min_row + amount,
            end_row=max_row + amount,
            start_column=min_col,
            end_column=max_col,
        )


def build_main_sheet_remark(match: dict[str, Any]) -> str:
    fit_status = normalize_fit_status(match.get("fit_status"))
    evidence_and_gaps = list(dict.fromkeys([*(match.get("missing_evidence", []) or []), *(match.get("gap_clauses", []) or [])]))
    if evidence_and_gaps:
        summary = "；".join(clean_text(item) for item in evidence_and_gaps[:2] if clean_text(item))
    else:
        summary = "；".join(clean_text(item) for item in str(match.get("reason") or "").split("；")[:2] if clean_text(item))
    return f"【{fit_status}】{summary}" if summary else f"【{fit_status}】"


def write_candidate_row(ws, row_number: int, section: dict[str, Any], match: dict[str, Any], sequence_number: int) -> None:
    copy_row_style(ws, section["first_data_row"], row_number, MAIN_SHEET_MAX_COL)
    columns = section["columns"]
    key_conditions = format_policy_conditions(match.get("关键申报条件") or "")
    remark = build_main_sheet_remark(match)
    values = {
        "序号": sequence_number,
        "申报部委": match.get("部委") or "",
        "项目名称": match.get("项目名称") or "",
        "申报时间（预估）": "",
        "项目性质": match.get("项目类型") or "",
        "资助金额": "",
        "关键申报条件": key_conditions,
        "备注": remark,
    }
    for header, col in columns.items():
        cell = ws.cell(row_number, col)
        cell.value = values.get(header, "")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row_number, columns["序号"]).fill = copy(AI_SEQUENCE_FILL)
    ws.row_dimensions[row_number].height = estimate_row_height(
        [
            (values.get("项目名称"), 18),
            (values.get("关键申报条件"), 26),
            (values.get("备注"), 20),
        ],
        min_height=34.15,
        line_height=17.0,
    )


def append_main_sheet_candidates(
    ws,
    section: dict[str, Any],
    matches: list[dict[str, Any]],
) -> None:
    if not matches:
        return

    used_rows = section_used_rows(ws, section)
    last_used_row = max(used_rows, default=section["header_row"])
    insertion_start = max(section["first_data_row"], last_used_row + 1)
    available_rows = [
        row_number
        for row_number in range(insertion_start, section["data_end_row"] + 1)
        if is_blank_main_sheet_row(ws, row_number)
    ]
    needed_rows = len(matches)
    if len(available_rows) < needed_rows:
        missing_rows = needed_rows - len(available_rows)
        insertion_row = section["data_end_row"] + 1
        shift_merged_ranges(ws, insertion_row, missing_rows)
        available_rows.extend(range(insertion_row, insertion_row + missing_rows))

    sequence_number = max_section_sequence(ws, section)
    for row_number, match in zip(available_rows[:len(matches)], matches):
        sequence_number += 1
        write_candidate_row(ws, row_number, section, match, sequence_number)


def write_main_sheet_candidates(workbook, matches: dict[str, Any], template_profile: dict[str, Any]) -> None:
    ws = workbook[template_profile["sheet_name"]]
    sections = scan_sheet_sections(ws, template_profile)
    candidates_by_section = build_main_sheet_candidates(matches, template_profile)
    existing_names = extract_existing_project_names_from_sheet(ws, template_profile)

    for section in sorted(sections, key=lambda item: item["title_row"], reverse=True):
        filtered_matches = []
        for match in candidates_by_section.get(section["name"], []):
            project_name = clean_text(match.get("项目名称", ""))
            if not project_name:
                continue
            normalized = normalize_string(project_name)
            if normalized in existing_names:
                continue
            existing_names.add(normalized)
            filtered_matches.append(match)
        append_main_sheet_candidates(ws, section, filtered_matches)


def write_review_sheet(workbook, profile: dict[str, Any], matches: dict[str, Any], template_profile: dict[str, Any], seed: dict[str, Any]) -> None:
    sheet_name = template_profile["review_sheet"]["name"]
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)

    ws["A1"] = f"企业：{seed.get('enterprise_name') or matches.get('company', {}).get('enterprise_name') or ''}"
    ws["A1"].font = Font(bold=True)
    row_number = 3
    row_number = write_summary_section(ws, row_number, "访谈纪要总结", profile.get("minutes_summary", {}))
    row_number += 1

    headers = template_profile["review_sheet"]["headers"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row_number, col)
        cell.value = header
        cell.font = Font(bold=True)
    row_number += 1
    for row in build_missing_rows(profile, template_profile):
        for col, value in enumerate(row, start=1):
            ws.cell(row_number, col).value = value
        row_number += 1

    row_number += 2
    ws.cell(row_number, 1).value = "上下文冲突"
    ws.cell(row_number, 1).font = Font(bold=True)
    row_number += 1
    conflict_headers = ["字段", "Excel 值", "纪要值", "处理方式", "说明"]
    for col, header in enumerate(conflict_headers, start=1):
        cell = ws.cell(row_number, col)
        cell.value = header
        cell.font = Font(bold=True)
    row_number += 1

    for row in build_conflict_rows(profile, template_profile):
        for col, value in enumerate(row, start=1):
            ws.cell(row_number, col).value = value
        row_number += 1

    row_number += 2
    ws.cell(row_number, 1).value = "匹配政策结果"
    ws.cell(row_number, 1).font = Font(bold=True)
    row_number += 1
    second_headers = ["区块", "项目名称", "匹配状态", "匹配理由", "缺失证据/待达条件", "来源行号"]
    for col, header in enumerate(second_headers, start=1):
        cell = ws.cell(row_number, col)
        cell.value = header
        cell.font = Font(bold=True)
    row_number += 1

    for row in build_policy_result_rows(matches):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row_number, col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_number += 1

    row_number += 2
    row_number = write_summary_section(ws, row_number, "匹配结果总结", matches.get("match_summary", {}))
    row_number += 1

    ws.cell(row_number, 1).value = "风险提示"
    ws.cell(row_number, 1).font = Font(bold=True)
    row_number += 1
    risk_headers = ["字段", "严重级别", "原因"]
    for col, header in enumerate(risk_headers, start=1):
        cell = ws.cell(row_number, col)
        cell.value = header
        cell.font = Font(bold=True)
    row_number += 1
    for item in profile.get("risk_flags", []):
        ws.cell(row_number, 1).value = item.get("field")
        ws.cell(row_number, 2).value = item.get("severity")
        ws.cell(row_number, 3).value = item.get("reason")
        row_number += 1

    for width_col, width in {"A": 20, "B": 42, "C": 14, "D": 44, "E": 30, "F": 12, "G": 12}.items():
        ws.column_dimensions[width_col].width = width


def main() -> int:
    args = parse_args()
    template_profile = resolve_template_profile(args.workbook, args.registry)
    seed = extract_seed_context(args.workbook, template_profile)
    matches = json.loads(args.matches.read_text(encoding="utf-8"))
    profile = json.loads(args.profile.read_text(encoding="utf-8"))

    workbook = load_workbook(args.workbook, data_only=False)
    try:
        write_main_sheet_candidates(workbook, matches, template_profile)
        write_review_sheet(workbook, profile, matches, template_profile, seed)
        args.output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(args.output)
    finally:
        workbook.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
