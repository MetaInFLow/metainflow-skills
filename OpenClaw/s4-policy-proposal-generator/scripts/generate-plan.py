#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
政策申报规划表生成脚本

功能：
1. 基于空表模板生成单 sheet 规划表
2. 支持仅填充企业信息，或填充企业信息 + 资质认定类/资助类项目
3. 输出文件统一规范到当前项目的 output/ 目录，文件名自动添加公司名前缀

调用方式：
  方式1（仅填企业信息）：
    python3 generate-plan.py --fill-company-only --company '{...}' --output 项目申报规划表.xlsx

  方式2（完整模式）：
    python3 generate-plan.py --company '{...}' --projects '[...]' --output 项目申报规划表.xlsx
"""

import argparse
import json
import math
import os
import re
import sys
import unicodedata
from copy import copy, deepcopy

from openpyxl import load_workbook


SECTION_LABELS = (
    "申报对象：",
    "申报条件：",
    "基础条件：",
    "专项条件：",
    "基本条件：",
    "认定标准：",
    "优先条件：",
    "优先范围：",
    "受理范围：",
    "受理标准：",
    "服务对象：",
    "支持对象：",
    "支持标准：",
    "奖励标准：",
)

SECTION_HEADERS = (
    "序号",
    "申报部委",
    "项目名称",
    "申报时间\n（预估）",
    "项目性质",
    "资助金额",
    "关键申报条件",
    "备注",
)

ALLOWED_PROJECT_CATEGORIES = ("qualification", "funding")
DEFAULT_TEMPLATE_FILE = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "references", "公司项目规划表_空表.xlsx")
)
DEFAULT_OUTPUT_FILENAME = "项目申报规划表.xlsx"
QUALIFICATION_TITLE = "①产业政策-资质认定类"
FUNDING_TITLE = "②产业政策-资助类"
QUALIFICATION_TITLE_ROW = 8
QUALIFICATION_HEADER_ROW = 9
QUALIFICATION_DATA_ROW = 10
FUNDING_TITLE_TEMPLATE_ROW = 11
SECTION_COLUMN_COUNT = 8

COMPANY_TEMPLATE_CELLS = {
    "company_name": "C2",
    "registered_capital": "C3",
    "registration_date": "C4",
    "employee_count": "C5",
    "declared_projects": "C6",
    "main_business": "C7",
    "existing_qualifications": "G2",
    "registered_address": "G3",
    "intellectual_property": "G4",
    "industry": "G5",
}

REQUIRED_COMPANY_KEYS = tuple(COMPANY_TEMPLATE_CELLS.keys())

MOCK_COMPANY_INFO = {
    "registered_capital": "100万元（示例）",
    "registration_date": "2024-01-01（示例）",
    "employee_count": "20人（示例）",
    "declared_projects": "暂无（示例）",
    "main_business": "AI智能体；企业知识库；自动化工具（示例）",
    "existing_qualifications": "暂无（示例）",
    "registered_address": "深圳市南山区（示例）",
    "intellectual_property": "商标3项；软著2项（示例）",
    "industry": "人工智能 / 软件与信息服务业（示例）",
}


def _normalize_multiline_text(value):
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").replace("\u3000", " ")
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def _format_policy_conditions(value):
    text = _normalize_multiline_text(value)
    if not text:
        return ""

    for label in sorted(SECTION_LABELS, key=len, reverse=True):
        text = re.sub(rf"(?<!\n){re.escape(label)}", f"\n{label}", text)

    text = re.sub(r"(?<!\n)([一二三四五六七八九十]+、)", r"\n\1", text)
    text = re.sub(r"(?<!\n)([（(][一二三四五六七八九十]+[)）])", r"\n\1", text)
    text = re.sub(r"(?<!\n)([（(]\d+[)）])", r"\n\1", text)
    text = re.sub(r"(?<!\n)([①②③④⑤⑥⑦⑧⑨⑩])", r"\n\1", text)
    text = re.sub(r"(?<!\n)([一二三四五六七八九十]类。)", r"\n\1", text)
    text = re.sub(r"([：:；;。])\s*(\d+\.(?=[^\d]))", r"\1\n\2", text)
    text = re.sub(r"(?<!\n)\s+(\d+\.(?=[^\d]))", r"\n\1", text)

    if "\n" not in text and len(text) > 120:
        text = re.sub(r"；\s*", "；\n", text)

    return re.sub(r"\n{2,}", "\n", text).strip()


def _display_width(text):
    return sum(2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1 for char in str(text))


def _estimate_wrapped_lines(text, chars_per_line):
    value = _normalize_multiline_text(text)
    if not value:
        return 1
    total = 0
    for line in value.split("\n"):
        total += max(1, math.ceil(max(1, _display_width(line)) / max(1, chars_per_line)))
    return total


def _estimate_row_height(texts, min_height=24.0, line_height=16.5, padding=8.0, max_height=409.5):
    line_count = 1
    for text, chars_per_line in texts:
        line_count = max(line_count, _estimate_wrapped_lines(text, chars_per_line))
    return round(min(max_height, max(min_height, line_count * line_height + padding)), 2)


def _parse_json_input(payload, expected_type, field_name, default=None):
    if payload is None:
        return deepcopy(default)
    value = json.loads(payload) if isinstance(payload, str) else payload
    if not isinstance(value, expected_type):
        type_name = "JSON对象" if expected_type is dict else "JSON数组"
        raise ValueError(f"{field_name}必须是{type_name}")
    return value


def _is_blank(value):
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    return False


def _stringify_value(value, list_joiner="；"):
    if _is_blank(value):
        return ""
    if isinstance(value, list):
        items = [str(item).strip() for item in value if not _is_blank(item)]
        return list_joiner.join(items)
    return str(value).strip()


def _company_input_count(company_info):
    return sum(1 for key in REQUIRED_COMPANY_KEYS if not _is_blank(company_info.get(key)))


def _prepare_company_info(company_info, fill_company_only=False):
    if not isinstance(company_info, dict):
        raise ValueError("企业信息必须是JSON对象")

    prepared = deepcopy(company_info)
    company_name = _stringify_value(prepared.get("company_name"))
    if not company_name:
        raise ValueError("企业名称不能为空")
    prepared["company_name"] = company_name

    auto_demo = fill_company_only and _company_input_count(prepared) == 1
    is_demo = bool(prepared.get("is_demo")) or auto_demo
    prepared["is_demo"] = is_demo

    for key in REQUIRED_COMPANY_KEYS:
        if key == "company_name":
            continue
        current = _stringify_value(prepared.get(key))
        if current:
            prepared[key] = current
            continue
        prepared[key] = MOCK_COMPANY_INFO[key] if is_demo else "待补充"

    return prepared


def _prepare_projects(projects):
    prepared = []
    for index, project in enumerate(projects, start=1):
        if not isinstance(project, dict):
            raise ValueError(f"项目列表第{index}项必须是JSON对象")

        item = deepcopy(project)
        category = _stringify_value(item.get("category"))
        if category not in ALLOWED_PROJECT_CATEGORIES:
            received = category or "空值"
            raise ValueError(
                f"项目列表第{index}项的category仅支持 qualification/funding，收到: {received}"
            )
        item["category"] = category
        prepared.append(item)
    return prepared


def _sanitize_filename_part(value):
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", _stringify_value(value, list_joiner="-"))
    cleaned = cleaned.strip(" .")
    return cleaned or "未命名公司"


def _prefix_output_filename(filename, company_name):
    basename = filename or DEFAULT_OUTPUT_FILENAME
    prefix = f"{_sanitize_filename_part(company_name)}-"
    return basename if basename.startswith(prefix) else f"{prefix}{basename}"


def _resolve_output_path(output_path, company_name):
    project_root = os.path.abspath(os.getcwd())
    project_output_dir = os.path.join(project_root, "output")
    normalized = os.path.normpath(output_path or DEFAULT_OUTPUT_FILENAME)
    basename = os.path.basename(normalized)
    dirname = os.path.dirname(normalized)

    if basename in {"", ".", os.curdir}:
        basename = DEFAULT_OUTPUT_FILENAME

    prefixed_basename = _prefix_output_filename(basename, company_name)

    if os.path.isabs(normalized):
        try:
            within_output_dir = os.path.commonpath([project_output_dir, normalized]) == project_output_dir
        except ValueError:
            within_output_dir = False
        target_dir = os.path.dirname(normalized) if within_output_dir else project_output_dir
        return os.path.join(target_dir, prefixed_basename)

    first_segment = normalized.split(os.sep, 1)[0]
    if first_segment == "output":
        target_dir = os.path.join(project_root, dirname)
    else:
        target_dir = os.path.join(project_output_dir, dirname)
    return os.path.join(target_dir, prefixed_basename)


def _load_template_sheet(template_file=None):
    template_path = os.path.normpath(template_file or DEFAULT_TEMPLATE_FILE)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]
    return workbook, worksheet


def _fill_company_template(ws, company_info):
    title_suffix = "（示例数据）" if company_info.get("is_demo") else ""
    ws["A1"] = f"{company_info['company_name']}\n  项目申报规划表{title_suffix}"
    ws.title = re.sub(r'[\\/*?:\[\]]', "_", company_info["company_name"])[:31] or "项目申报规划表"

    for key, cell_ref in COMPANY_TEMPLATE_CELLS.items():
        ws[cell_ref] = company_info.get(key, "")


def _unmerge_range_if_needed(ws, cell_range):
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == cell_range:
            try:
                ws.unmerge_cells(cell_range)
            except KeyError:
                # openpyxl 在插行后可能保留失效的 merge 元数据，直接移除即可。
                ws.merged_cells.ranges.discard(merged_range)
            return


def _copy_row_style(ws, source_row, target_row):
    for column in range(1, SECTION_COLUMN_COUNT + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        target._style = copy(source._style)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _snapshot_row_style(ws, row_index):
    return {
        "styles": [copy(ws.cell(row_index, column)._style) for column in range(1, SECTION_COLUMN_COUNT + 1)],
        "height": ws.row_dimensions[row_index].height,
    }


def _apply_row_style_snapshot(ws, row_style_snapshot, target_row):
    for column, style in enumerate(row_style_snapshot["styles"], start=1):
        ws.cell(target_row, column)._style = copy(style)
    ws.row_dimensions[target_row].height = row_style_snapshot["height"]


def _apply_section_scaffold(ws, title_row, header_row, title):
    title_range = f"A{title_row}:H{title_row}"
    _unmerge_range_if_needed(ws, title_range)
    ws.merge_cells(title_range)
    ws.cell(title_row, 1).value = title

    for column, header in enumerate(SECTION_HEADERS, start=1):
        ws.cell(header_row, column).value = header


def _build_project_remark(project):
    parts = []
    match_reason = _normalize_multiline_text(_stringify_value(project.get("match_reason")))
    source_link = _normalize_multiline_text(_stringify_value(project.get("source_link")))

    if match_reason:
        parts.append(f"匹配理由：{match_reason}")
    if source_link:
        parts.append(f"来源链接：{source_link}")
    return "\n".join(parts)


def _write_placeholder_row(ws, row_index, message):
    placeholder_range = f"A{row_index}:H{row_index}"
    _unmerge_range_if_needed(ws, placeholder_range)
    ws.merge_cells(placeholder_range)
    ws.cell(row_index, 1).value = message
    ws.row_dimensions[row_index].height = 34.15


def _write_project_row(ws, row_index, project, serial_number):
    key_conditions = _format_policy_conditions(_stringify_value(project.get("key_conditions"), list_joiner="\n"))
    remark = _build_project_remark(project)
    values = (
        serial_number,
        _stringify_value(project.get("department")),
        _stringify_value(project.get("project_name")),
        _stringify_value(project.get("application_time")),
        _stringify_value(project.get("project_type")),
        _stringify_value(project.get("funding_amount")),
        key_conditions,
        remark,
    )

    for column, value in enumerate(values, start=1):
        ws.cell(row_index, column).value = value

    ws.row_dimensions[row_index].height = _estimate_row_height(
        [
            (values[2], 16),
            (key_conditions, 28),
            (remark, 22),
        ],
        min_height=34.15,
        line_height=17.0,
    )


def _populate_project_section(ws, start_row, projects):
    if not projects:
        _write_placeholder_row(ws, start_row, "暂无匹配项目")
        return

    for offset, project in enumerate(projects):
        _write_project_row(ws, start_row + offset, project, offset + 1)


def _populate_policy_sections(ws, projects):
    qualification_projects = [project for project in projects if project["category"] == "qualification"]
    funding_projects = [project for project in projects if project["category"] == "funding"]
    qualification_data_style = _snapshot_row_style(ws, QUALIFICATION_DATA_ROW)
    funding_title_style = _snapshot_row_style(ws, FUNDING_TITLE_TEMPLATE_ROW)
    funding_header_style = _snapshot_row_style(ws, FUNDING_TITLE_TEMPLATE_ROW + 1)
    funding_data_style = _snapshot_row_style(ws, FUNDING_TITLE_TEMPLATE_ROW + 2)

    if len(qualification_projects) > 1:
        _unmerge_range_if_needed(ws, "A11:H11")
        for target_row in range(
            QUALIFICATION_DATA_ROW + 1,
            QUALIFICATION_DATA_ROW + len(qualification_projects),
        ):
            _apply_row_style_snapshot(ws, qualification_data_style, target_row)

    funding_title_row = FUNDING_TITLE_TEMPLATE_ROW + max(0, len(qualification_projects) - 1)
    funding_header_row = funding_title_row + 1
    funding_data_row = funding_title_row + 2

    _apply_row_style_snapshot(ws, funding_title_style, funding_title_row)
    _apply_row_style_snapshot(ws, funding_header_style, funding_header_row)

    funding_row_count = max(1, len(funding_projects))
    for target_row in range(funding_data_row, funding_data_row + funding_row_count):
        _apply_row_style_snapshot(ws, funding_data_style, target_row)

    _apply_section_scaffold(ws, QUALIFICATION_TITLE_ROW, QUALIFICATION_HEADER_ROW, QUALIFICATION_TITLE)
    _apply_section_scaffold(ws, funding_title_row, funding_header_row, FUNDING_TITLE)
    _populate_project_section(ws, QUALIFICATION_DATA_ROW, qualification_projects)
    _populate_project_section(ws, funding_data_row, funding_projects)


def generate_company_template_plan(company_info_json, output_path, template_file=None):
    """
    基于空表模板填充企业信息并生成单 sheet 规划表。
    """
    company_info = _prepare_company_info(
        _parse_json_input(company_info_json, dict, "企业信息"),
        fill_company_only=True,
    )
    workbook, worksheet = _load_template_sheet(template_file)
    _fill_company_template(worksheet, company_info)

    final_output_path = _resolve_output_path(output_path, company_info["company_name"])
    output_dir = os.path.dirname(final_output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook.save(final_output_path)
    return final_output_path


def generate_plan(company_info_json, projects_json, output_path, template_file=None):
    """
    基于空表模板生成包含企业信息、资质认定类和资助类项目的单 sheet 规划表。
    """
    company_info = _prepare_company_info(_parse_json_input(company_info_json, dict, "企业信息"))
    projects = _prepare_projects(_parse_json_input(projects_json, list, "项目列表", default=[]))

    workbook, worksheet = _load_template_sheet(template_file)
    _fill_company_template(worksheet, company_info)
    _populate_policy_sections(worksheet, projects)

    final_output_path = _resolve_output_path(output_path, company_info["company_name"])
    output_dir = os.path.dirname(final_output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook.save(final_output_path)
    return final_output_path


def main():
    """主函数，支持命令行调用。"""
    parser = argparse.ArgumentParser(
        description="生成政策申报规划表Excel文件",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 方式1：仅填充模板顶部企业信息
  python3 generate-plan.py --fill-company-only --company '{"company_name":"测试公司"}' --output 项目申报规划表.xlsx

  # 方式2：传入企业信息和项目列表，输出完整单sheet规划表
  python3 generate-plan.py --company '{"company_name":"测试公司",...}' --projects '[{"category":"qualification",...}]' --output 项目申报规划表.xlsx

  # 方式3：从文件读取JSON
  python3 generate-plan.py --company-file company.json --projects-file projects.json --output 项目申报规划表.xlsx
        """,
    )

    parser.add_argument("--company", type=str, help="企业信息JSON字符串")
    parser.add_argument("--projects", type=str, help="项目列表JSON字符串")
    parser.add_argument("--fill-company-only", action="store_true", help="仅填充空表模板中的企业信息区")
    parser.add_argument("--template-file", type=str, default=DEFAULT_TEMPLATE_FILE, help="空表模板路径")
    parser.add_argument("--company-file", type=str, help="企业信息JSON文件路径")
    parser.add_argument("--projects-file", type=str, help="项目列表JSON文件路径")
    parser.add_argument(
        "--output",
        type=str,
        default=DEFAULT_OUTPUT_FILENAME,
        help="输出文件名或路径；最终会规范到当前项目的 output/ 目录，并自动添加公司名前缀",
    )

    args = parser.parse_args()

    if args.company:
        company_info_json = args.company
    elif args.company_file:
        with open(args.company_file, "r", encoding="utf-8") as file:
            company_info_json = file.read()
    else:
        print("错误：请通过 --company 或 --company-file 提供企业信息")
        sys.exit(1)

    projects_json = None
    if not args.fill_company_only:
        if args.projects:
            projects_json = args.projects
        elif args.projects_file:
            with open(args.projects_file, "r", encoding="utf-8") as file:
                projects_json = file.read()
        else:
            print("错误：请通过 --projects 或 --projects-file 提供项目列表")
            sys.exit(1)

    try:
        if args.fill_company_only:
            result = generate_company_template_plan(
                company_info_json,
                args.output,
                template_file=args.template_file,
            )
        else:
            result = generate_plan(
                company_info_json,
                projects_json,
                args.output,
                template_file=args.template_file,
            )
        print(f"规划表生成成功: {result}")
    except Exception as exc:
        print(f"生成失败: {str(exc)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
